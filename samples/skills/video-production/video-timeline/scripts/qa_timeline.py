#!/usr/bin/env python3
"""qa_timeline.py \u2014 objective 100-point QA score for a story-timeline xlsx.

Part of the `video-timeline` skill. The 12 checks, weights, and thresholds
mirror references/timing-best-practices.md section 4 EXACTLY \u2014 if you change
one, change both. Prints a JSON verdict: score, grade, per-check
score/weight/evidence, hard fails, and a prioritized fix list.

Grades: A >= 90, B >= 75, C >= 60, D < 60.
Definition of done: grade >= B AND zero hard fails.

Usage:
  python3 qa_timeline.py --timeline <name>.xlsx \\
      --transcript <narration.json> [--words <transcript.txt>]
"""
import argparse
import json
import sys

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl is required: pip install openpyxl")

VALID_ROLES = {"Hook", "Promise", "Establish", "Story", "Proof", "Re-hook",
               "Payoff", "Reveal", "Principle", "Transition", "CTA", "Close"}
VALID_TYPES = {"New Image", "Slide", "Video Clip", "Session B-roll", "End Card", "Hold"}
VALID_STATUS = {"To-do", "Have", "Verify", "Optional", "Done"}
TOMAKE = {"New Image", "Slide", "Video Clip"}


def t2s(t):
    try:
        m, s = str(t).split(":")
        return int(m) * 60 + float(s)
    except Exception:
        return None


def load_rows(path):
    ws = openpyxl.load_workbook(path, data_only=True).active
    rows = []
    for r in range(2, ws.max_row + 1):
        i = ws.cell(r, 1).value
        if i is None:
            continue
        in_t = t2s(ws.cell(r, 4).value)
        out_t = t2s(ws.cell(r, 5).value)
        if in_t is None or out_t is None:
            continue
        rows.append({
            "id": int(i),
            "chapter": ws.cell(r, 2).value or "",
            "moment": (ws.cell(r, 3).value or "").strip(),
            "in": in_t, "out": out_t,
            "says": (ws.cell(r, 7).value or "").strip(),
            "role": (ws.cell(r, 8).value or "").strip(),
            "type": (ws.cell(r, 9).value or "").strip(),
            "show": (ws.cell(r, 10).value or "").strip(),
            "asset": (ws.cell(r, 11).value or "").strip(),
            "status": (ws.cell(r, 12).value or "").strip(),
            "notes": (ws.cell(r, 13).value or "").strip(),
        })
    return rows


def grade(score):
    return "A" if score >= 90 else "B" if score >= 75 else "C" if score >= 60 else "D"


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--timeline", required=True)
    ap.add_argument("--transcript", required=True, help="narration.json for audio duration")
    ap.add_argument("--words", help="word-timed transcript (informational)")
    args = ap.parse_args()

    rows = load_rows(args.timeline)
    dur = float(json.load(open(args.transcript, encoding="utf-8")).get("duration") or 0)
    checks = []

    def add(no, name, weight, passed, total, evidence, hard=False):
        frac = 1.0 if total == 0 else passed / total
        score = round(weight * frac, 1)
        checks.append({
            "check": no, "name": name, "weight": weight, "score": score,
            "pass": "PASS" if frac >= 0.999 else ("PARTIAL" if frac > 0 else "FAIL"),
            "evidence": evidence, "hard": hard,
        })

    if not rows:
        print(json.dumps({"error": "no rows found in timeline (check the Timeline sheet)"}))
        sys.exit(1)

    # --- 1. Chain integrity (HARD, 15) -----------------------------------
    bad = []
    for a, b in zip(rows, rows[1:]):
        if b["in"] < a["out"] - 0.5:
            bad.append("overlap rows %d/%d" % (a["id"], b["id"]))
        elif b["in"] > a["out"] + 2.5:
            bad.append("gap %.1fs after row %d" % (b["in"] - a["out"], a["id"]))
    if rows[0]["in"] > 2.5:
        bad.append("does not start at 0:00 (starts %.1fs)" % rows[0]["in"])
    if rows[-1]["out"] < dur - 0.5:
        bad.append("coverage ends %.1fs, audio ends %.1fs" % (rows[-1]["out"], dur))
    add(1, "chain integrity", 15, len(rows) - len(bad), len(rows),
        "; ".join(bad[:6]) if bad else "unbroken chain 0:00 -> %.1fs" % rows[-1]["out"], hard=True)

    # --- 2. Tail pad (HARD, 5) -------------------------------------------
    last = rows[-1]
    ok = last["out"] - dur >= 1.5 and last["type"] in ("End Card", "Hold")
    add(2, "tail pad", 5, 1 if ok else 0, 1,
        "last row ends %.1fs (%.1fs past audio), type=%s" % (last["out"], last["out"] - dur, last["type"] or "?"), hard=True)

    # --- 3. No 0-s rows (HARD, 5) -----------------------------------------
    zero = [r["id"] for r in rows if r["out"] - r["in"] <= 0]
    add(3, "no 0-s rows", 5, len(rows) - len(zero), len(rows),
        "0-s rows: %s" % zero if zero else "all rows >0s", hard=True)

    # --- 4. No oversized statics (10) -------------------------------------
    over18 = [r["id"] for r in rows if r["out"] - r["in"] > 18]
    over30 = [r["id"] for r in rows if r["out"] - r["in"] > 30]
    frac = (len(rows) - len(over18)) / len(rows)
    if over30:
        frac = min(frac, 0.5)
    add(4, "no oversized statics", 10, int(round(frac * len(rows))), len(rows),
        "%d rows >18s (%s); %d rows >30s (%s)" % (len(over18), over18[:8], len(over30), over30[:8]))

    # --- 5. Long-run split (10) -------------------------------------------
    problems = []
    # detect consecutive Session B-roll / Video Clip stretches
    i = 0
    while i < len(rows):
        j = i
        while j + 1 < len(rows) and rows[j + 1]["in"] <= rows[j]["out"] + 0.01 and rows[j + 1]["type"] in ("Session B-roll", "Video Clip"):
            j += 1
        span = rows[j]["out"] - rows[i]["in"]
        if span > 60:
            has_hook = any(r["role"] in ("Re-hook", "Payoff", "Reveal") for r in rows[i:j + 1])
            if not has_hook and (j - i + 1) < 2:
                problems.append("rows %d-%d: %.0fs run of one visual" % (rows[i]["id"], rows[j]["id"], span))
        i = j + 1
    # also: single row >60s of any type without a hook role
    for r in rows:
        if r["out"] - r["in"] > 60 and r["role"] not in ("Re-hook", "Payoff", "Reveal", "CTA"):
            problems.append("row %d: %.0fs single visual" % (r["id"], r["out"] - r["in"]))
    problems = sorted(set(problems))
    add(5, "long-run split", 10, 1 if not problems else 0, 1,
        "; ".join(problems[:5]) if problems else "no >60s unbroken visual run without a re-hook")

    # --- 6. Text-beat cadence (10) ----------------------------------------
    text_rows = [r for r in rows if r["type"] in ("Slide", "New Image", "End Card") and r["show"]]
    covered = 0
    win = 0.0
    while win < dur:
        w1 = min(win + 75.0, dur)
        if any(r["in"] < w1 and r["out"] > win for r in text_rows):
            covered += 1
        win = w1
    total_w = max(1, int((dur + 74.9) // 75))
    add(6, "text-beat cadence (every 75s)", 10, covered, total_w,
        "%d/%d 75s windows contain a card/slide (text rows: %d)" % (covered, total_w, len(text_rows)))

    # --- 7. Opening (8) ----------------------------------------------------
    p7 = []
    hookish = [r for r in rows if r["role"] in ("Promise", "Hook") and r["in"] <= 30]
    p7.append(("promise row <=30s", bool(hookish)))
    p7.append(("first row is strong frame (New Image/Slide)", rows[0]["type"] in ("New Image", "Slide")))
    p7.append(("first row <=10s", rows[0]["out"] - rows[0]["in"] <= 10))
    add(7, "opening", 8, sum(1 for _, ok in p7 if ok), len(p7),
        "; ".join("%s=%s" % (n, "ok" if ok else "MISSING") for n, ok in p7))

    # --- 8. Act/chapter plan (8) ------------------------------------------
    acts = {}
    for r in rows:
        acts.setdefault(r["chapter"], [r["in"], r["out"]])
        acts[r["chapter"]][1] = max(acts[r["chapter"]][1], r["out"])
    alen = sorted(v[1] - v[0] for v in acts.values())
    ok8 = []
    ok8.append(("acts 4-15", 4 <= len(acts) <= 15))
    ok8.append(("act lengths 60-240s", all(60 <= a <= 240 for a in alen)))
    ok8.append(("first act <=90s", alen[0] <= 90 if alen else False))
    add(8, "act/chapter plan", 8, sum(1 for _, ok in ok8 if ok), len(ok8),
        "%d acts, lengths %s; " % (len(acts), alen) + "; ".join("%s=%s" % (n, "ok" if ok else "MISSING") for n, ok in ok8))

    # --- 9. Slump-zone pivot (8) -------------------------------------------
    lo, hi = 0.40 * dur, 0.55 * dur
    pivots = [r["id"] for r in rows if r["role"] in ("Re-hook", "Payoff", "Reveal", "CTA") and lo <= r["in"] <= hi]
    add(9, "slump-zone pivot (40-55%%)", 8, 1 if pivots else 0, 1,
        "pivot rows in %.0fs-%.0fs: %s" % (lo, hi, pivots if pivots else "NONE"))

    # --- 10. Ending (8) ------------------------------------------------------
    ok10 = []
    endcard = [r for r in rows if r["type"] == "End Card" and r["in"] >= dur - 20 - 5]
    ok10.append(("End Card in last ~20s", bool(endcard)))
    ctas = [r for r in rows if r["role"] == "CTA"]
    ok10.append(("exactly one CTA row", len(ctas) == 1))
    ok10.append(("CTA near end (>= dur-40s)", bool(ctas) and ctas[0]["in"] >= dur - 40))
    add(10, "ending", 8, sum(1 for _, ok in ok10 if ok), len(ok10),
        "; ".join("%s=%s" % (n, "ok" if ok else "MISSING") for n, ok in ok10))

    # --- 11. Fills complete (7) ---------------------------------------------
    filled = sum(1 for r in rows if r["show"])
    short = sum(1 for r in rows if len(r["says"].split()) <= 60 and r["says"])
    add(11, "fills complete", 7, int(round(10 * (0.5 * filled / len(rows) + 0.5 * short / len(rows)))), 10,
        "What-to-show %d/%d; plain-text<=60w %d/%d" % (filled, len(rows), short, len(rows)))

    # --- 12. Status hygiene (6) ---------------------------------------------
    badst = [r["id"] for r in rows if r["status"] not in VALID_STATUS]
    badtype = [r["id"] for r in rows if r["type"] not in VALID_TYPES]
    done_no_asset = [r["id"] for r in rows if r["status"] == "Done" and r["type"] in TOMAKE and not r["asset"]]
    badrole = [r["id"] for r in rows if r["role"] not in VALID_ROLES]
    bad12 = badst + badtype + badrole
    ok12 = not bad12 and not done_no_asset
    add(12, "status hygiene", 6, 1 if ok12 else 0, 1,
        "; ".join(filter(None, [
            "invalid status: %s" % badst if badst else "",
            "invalid type: %s" % badtype if badtype else "",
            "invalid role: %s" % badrole if badrole else "",
            "Done w/o asset (to-make types): %s" % done_no_asset if done_no_asset else "",
        ])) or "all valid")

    total = round(sum(c["score"] for c in checks), 1)
    hard_fails = [c for c in checks if c["hard"] and c["pass"] != "PASS"]
    fixes = [c for c in sorted(checks, key=lambda c: c["score"] / c["weight"])
             if c["pass"] != "PASS"][:5]

    out = {
        "timeline": args.timeline,
        "audio_duration_s": dur,
        "rows": len(rows),
        "score": total,
        "grade": grade(total),
        "definition_of_done": "grade>=B and no hard fails",
        "done": (grade(total) in ("A", "B")) and not hard_fails,
        "hard_fails": [c["name"] + ": " + c["evidence"] for c in hard_fails],
        "checks": checks,
        "fix_next": [c["name"] + " -> " + c["evidence"] for c in fixes],
    }
    print(json.dumps(out, indent=1))


if __name__ == "__main__":
    main()
