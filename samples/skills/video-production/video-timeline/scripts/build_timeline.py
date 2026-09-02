#!/usr/bin/env python3
"""build_timeline.py \u2014 generate the editable story-timeline xlsx from a final
TTS narration (timing JSON + word-timed transcript).

Part of the `video-timeline` skill (video-production set). See SKILL.md and
references/timeline-format.md for the schema and the assistant pass that
follows generation.

Stdlib + openpyxl. Read-only on inputs; writes the xlsx (and prints a JSON
summary to stdout).
"""
import argparse
import json
import os
import re
import sys

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("openpyxl is required: pip install openpyxl")


def tcs(s):
    s = int(round(s))
    return "%d:%02d" % divmod(s, 60)


def load_segments(path):
    d = json.load(open(path, encoding="utf-8"))
    segs = d.get("segments") or []
    segs = [(float(s["start"]), float(s["end"]), str(s.get("text", ""))) for s in segs]
    segs = [s for s in segs if s[1] > s[0]]  # drop 0-s TTS seam segments
    segs.sort()
    dur = float(d.get("duration") or 0) or (segs[-1][1] if segs else 0.0)
    return segs, dur


def load_words(path):
    """word-timed transcript -> [(start, end, word)]; None if absent."""
    if not path or not os.path.exists(path):
        return None
    words = []
    pat = re.compile(r"\[\s*([\d.]+)\s*->\s*([\d.]+)\]\s*(\S+)")
    for line in open(path, encoding="utf-8"):
        m = pat.search(line)
        if m:
            words.append((float(m.group(1)), float(m.group(2)), m.group(3)))
    return words or None


def gaps_from_words(words, min_gap):
    """candidate break points: times of word gaps >= min_gap."""
    out = []
    for (a0, a1, _), (b0, b1, _) in zip(words, words[1:]):
        if b0 - a1 >= min_gap:
            out.append(a1)
    return out


def derive_beats(segs, words, target_max=18.0, target_min=8.0, gap_break=0.6):
    """Derive story beats (rows) from segments + word gaps.

    Break at a segment boundary when:
      - the current beat already reaches `target_max`, or
      - a real word gap (>= `gap_break`) sits just before the boundary and
        the current beat is already >= `target_min`.
    Sliver beats (< 2 s) are merged into the previous beat.
    """
    breaks = gaps_from_words(words, gap_break) if words else []
    beats = []
    cur = 0.0
    for s, e, _ in segs:
        if beats:
            if s - cur >= target_max:
                beats.append((cur, s))
                cur = s
            elif s - cur >= target_min and any(abs(s - b) < 1.5 for b in breaks):
                beats.append((cur, s))
                cur = s
    last_end = segs[-1][1] if segs else 0.0
    if cur < last_end:
        beats.append((cur, last_end))
    merged = []
    for b in beats:
        if b[1] - b[0] < 2.0 and merged:
            merged[-1] = (merged[-1][0], b[1])
        else:
            merged.append(b)
    if len(merged) > 1 and merged[0][1] - merged[0][0] < 2.0:
        merged[0] = (merged[0][0], merged[1][1])
        del merged[1]
    return merged


def load_beats_file(path):
    """Accept: list of [start,end] pairs; or {"beats": {...}} / {"beats": [...]}.
    Returns [(start,end),...]."""
    d = json.load(open(path, encoding="utf-8"))
    items = d if isinstance(d, list) else d.get("beats")
    if isinstance(items, dict):
        items = list(items.values())
    out = []
    for it in items or []:
        if isinstance(it, (list, tuple)):
            s, e = float(it[0]), float(it[1])
        elif isinstance(it, dict):
            s = float(it.get("start", it.get("fs", it.get("t0"))))
            e = float(it.get("end", it.get("fe", it.get("t1"))))
        else:
            continue
        if e >= s:
            out.append((s, e))
    out.sort()
    return out


def make_acts(rows):
    """Group rows into acts: first act 30-90 s, later acts 60-240 s
    (force-end at 210 s). Returns [(act_no, start, end), ...]."""
    if not rows:
        return []
    acts = []
    i = 0
    n = len(rows)
    while i < n:
        start = rows[i][0]
        limit = 90.0 if not acts else 210.0
        minlen = 30.0 if not acts else 60.0
        j = i
        while j + 1 < n and (rows[j][1] - start) < limit and (rows[j + 1][0] - start) < limit:
            j += 1
        if rows[j][1] - start >= limit:
            j -= 1
        acts.append((len(acts) + 1, start, rows[j][1]))
        i = j + 1
    return acts


def audio_levels(wav_path):
    """Streaming peak/RMS: 0.5 s chunks into array('h') (~48 KB live memory).
    Never materialize the whole sample buffer as Python objects."""
    import wave, math, array
    out = {}
    wf = wave.open(wav_path)
    sw = wf.getsampwidth()
    rate = wf.getframerate()
    peak = 0
    sq = 0.0
    total = 0
    if sw == 2:
        while True:
            raw = wf.readframes(max(1, rate // 2))
            if not raw:
                break
            buf = array.array("h")
            buf.frombytes(raw)
            cmax = 0
            csq = 0
            for v in buf:
                if v < 0:
                    v = -v
                if v > cmax:
                    cmax = v
                csq += v * v
            if cmax > peak:
                peak = cmax
            sq += csq
            total += len(buf)
    wf.close()
    if total:
        peak /= 32768.0
        rms = math.sqrt(sq / total) / 32768.0
        out["peak_dbfs"] = round(20 * math.log10(peak), 1) if peak > 0 else -99
        out["rms_dbfs"] = round(20 * math.log10(rms), 1) if rms > 0 else -99
    return out


def audit(segs, dur, words, wav_path=None):
    """Phase-1 audio audit. Returns a dict of numbers for the Audit sheet."""
    a = {"duration_s": round(dur, 1), "duration_mss": tcs(dur)}
    if segs:
        text = " ".join(s[2] for s in segs)
        a["word_count"] = len(text.split())
        a["wpm_overall"] = round(a["word_count"] / (dur / 60.0), 1) if dur else None
        if dur:
            cur = 0.0
            buckets = {}
            while cur < dur:
                b0, b1 = cur, min(cur + 30.0, dur)
                w = sum(len(t.split()) for s, e, t in segs if b0 <= s < b1)
                buckets[int(b0)] = round(w / (b1 - b0) * 60.0, 1)
                cur = b1
            if buckets:
                a["wpm_30s_min"] = min(buckets.values())
                a["wpm_30s_max"] = max(buckets.values())
                a["wpm_30s_min_at"] = tcs(min(buckets, key=buckets.get))
                a["wpm_30s_max_at"] = tcs(max(buckets, key=buckets.get))
    a["gaps_over_05"] = a["gaps_over_10"] = a["gaps_over_20"] = 0
    top = []
    if words:
        for (a0, a1, wa), (b0, b1, wb) in zip(words, words[1:]):
            g = b0 - a1
            if g > 0.5:
                a["gaps_over_05"] += 1
            if g > 1.0:
                a["gaps_over_10"] += 1
            if g > 2.0:
                a["gaps_over_20"] += 1
            if g > 0.6:
                top.append((g, a1, wa, wb))
        top.sort(reverse=True)
        a["top_pauses"] = ["%.2fs @ %s (..%s > %s..)" % (g, tcs(t), w1, w2) for g, t, w1, w2 in top[:8]]
        a["first_word_start"] = round(words[0][0], 2)
        a["last_word_end"] = round(words[-1][1], 2)
        a["tail_silence"] = round(dur - words[-1][1], 2)
    if wav_path and os.path.exists(wav_path):
        try:
            a.update(audio_levels(wav_path))
        except Exception as e:
            a["wav_error"] = str(e)
    return a


ROLE_COLORS = {
    "Hook": "FDE9C9", "Promise": "D6E4F5", "Establish": "EDEDED",
    "Story": "FFFFFF", "Proof": "DFF0E5", "Re-hook": "F5D6D0",
    "Payoff": "FCE4B0", "Reveal": "FCE4B0", "Principle": "EDEDED",
    "Transition": "EDEDED", "CTA": "D6E4F5", "Close": "EDEDED",
}
VTYPE = ["New Image", "Slide", "Video Clip", "Session B-roll", "End Card", "Hold"]
STATUS = ["To-do", "Have", "Verify", "Optional", "Done"]
ROLES = list(ROLE_COLORS.keys())


def words60(text):
    w = text.split()
    return " ".join(w[:60])


def build_xlsx(path, rows, acts, audit_res, dur, seam_notes):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Timeline"
    headers = ["#", "Chapter", "Moment", "In", "Out", "Dur",
               "What is said here (plain)", "Story role", "Visual type",
               "What to show", "Asset / file (you fill)", "Status", "Notes"]
    ws.append(headers)

    act_of = {}
    for no, a0, a1 in acts:
        for r in rows:
            if a0 <= r[0] < a1 and r[0] not in act_of:
                act_of[r[0]] = no
    last_act = acts[-1][0] if acts else 1
    for r in rows:
        if r[0] not in act_of:
            act_of[r[0]] = last_act

    for i, (s, e) in enumerate(rows, 1):
        vt = "New Image" if i == 1 else ("End Card" if i == len(rows) else "")
        role = "Hook" if i == 1 else ("Close" if i == len(rows) else "Story")
        ws.append([i, "Act %d" % act_of[s], "", tcs(s), tcs(e), round(e - s, 1),
                   "", role, vt, "", "", "To-do",
                   ("merges 0-s seam" if i in seam_notes else "")])

    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill = PatternFill("solid", fgColor="181B1F")
    for c in ws[1]:
        c.fill = hdr_fill
        c.font = Font(bold=True, color="FFFFFF", size=11)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border
    ws.row_dimensions[1].height = 30
    for r in range(2, ws.max_row + 1):
        role = ws.cell(r, 8).value
        fill = ROLE_COLORS.get(role, "FFFFFF")
        for c in range(1, 14):
            cell = ws.cell(r, c)
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if c in (1, 4, 5, 6, 9, 12):
                cell.alignment = Alignment(vertical="top", horizontal="center")
            if c == 2:
                cell.fill = PatternFill("solid", fgColor=fill)
        ws.cell(r, 3).font = Font(bold=True)
        ws.cell(r, 9).font = Font(bold=True)
    for col, w in {1: 4, 2: 12, 3: 17, 4: 6, 5: 6, 6: 6, 7: 52, 8: 11, 9: 14,
                   10: 36, 11: 20, 12: 9, 13: 32}.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    dvv = DataValidation(type="list", formula1='"' + ",".join(VTYPE) + '"', allow_blank=True)
    dvs = DataValidation(type="list", formula1='"' + ",".join(STATUS) + '"', allow_blank=True)
    dvr = DataValidation(type="list", formula1='"' + ",".join(ROLES) + '"', allow_blank=True)
    ws.add_data_validation(dvv)
    ws.add_data_validation(dvs)
    ws.add_data_validation(dvr)
    last = ws.max_row
    dvv.add("I2:I%d" % last)
    dvs.add("L2:L%d" % last)
    dvr.add("H2:H%d" % last)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:M%d" % last

    a = wb.create_sheet("Available assets")
    a.append(["Type", "What it is", "Where / how many", "Use for"])
    a.append(["Session B-roll", "Real footage from the source screen recording",
              "the session video.mp4 \u2014 cut any span",
              "moments where real 'you at work' footage fits"])
    a.append(["Rendered cards", "any pre-rendered title/text cards from earlier passes",
              "list them here", "thesis, credentials, reveals, promises"])
    a.append(["Website / product screenshots", "captures of the real UI",
              "list them here", "establish shots, portability, 'what it looks like'"])
    for c in a[1]:
        c.fill = hdr_fill
        c.font = Font(bold=True, color="FFFFFF")
        c.border = border
    for r in range(2, a.max_row + 1):
        for c in range(1, 5):
            a.cell(r, c).border = border
            a.cell(r, c).alignment = Alignment(vertical="top", wrap_text=True)
    for col, w in {1: 26, 2: 46, 3: 44, 4: 44}.items():
        a.column_dimensions[get_column_letter(col)].width = w
    a.freeze_panes = "A2"

    h = wb.create_sheet("How to use")
    lines = [
        "HOW TO USE",
        "",
        "1. Each row is a story moment. Column G is a plain summary of what the fixed narration says there.",
        "2. Pick a Visual type (column I dropdown): New Image / Slide / Video Clip / Session B-roll / End Card / Hold.",
        "3. Fill 'Asset / file' (column K) with the image, slide, or clip you want.",
        "4. Move Status (column L): To-do -> Have -> Verify -> Done.",
        "5. Split or merge rows freely \u2014 keep the chain unbroken (row[i+1].In == row[i].Out).",
        "",
        "Roles: Hook (first 30 s) - Promise (what they get) - Establish (who/what/why) - Story (default)",
        "      Proof (credentials/scale) - Re-hook (re-grab) - Payoff/Reveal (the big gives)",
        "      Principle (a rule) - Transition (act change) - CTA (next step) - Close (sign-off)",
        "",
        "Types: New Image = fresh generated/sourced image. Slide = text card you build.",
        "      Video Clip = full clip you record/pull. Session B-roll = source recording footage.",
        "      End Card = closing card. Hold = keep the previous visual across a 0-1 s seam.",
    ]
    for l in lines:
        h.append([l])
    h.column_dimensions["A"].width = 118
    h["A1"].font = Font(bold=True, size=14)

    au = wb.create_sheet("Audit")
    au.append(["Metric", "Value", "Threshold", "Pass?"])
    wpm = audit_res.get("wpm_overall")
    checks = [
        ("duration", "%s (%ss)" % (audit_res.get("duration_mss"), audit_res.get("duration_s")),
         "how-to 1-5 / explainer 8-12 / training <6 min", "see content type"),
        ("overall WPM", wpm, "130-165 (explainer/presentation)",
         "PASS" if (wpm or 999) <= 170 else "rushed"),
        ("WPM 30-s min", "%s @ %s" % (audit_res.get("wpm_30s_min"), audit_res.get("wpm_30s_min_at")),
         "<=170", "PASS" if (audit_res.get("wpm_30s_min") or 99) <= 170 else "rushed zone(s)"),
        ("WPM 30-s max", "%s @ %s" % (audit_res.get("wpm_30s_max"), audit_res.get("wpm_30s_max_at")),
         "<=170", "PASS" if (audit_res.get("wpm_30s_max") or 99) <= 170 else "rushed"),
        ("gaps >0.5 / >1.0 / >2.0 s", "%s / %s / %s" % (audit_res.get("gaps_over_05"),
         audit_res.get("gaps_over_10"), audit_res.get("gaps_over_20")),
         "deliberate 1-3 s silences wanted",
         "thin" if (audit_res.get("gaps_over_10") or 0) < 5 else "ok"),
        ("tail silence", "%ss" % audit_res.get("tail_silence"), ">=0.5 s",
         "PASS" if (audit_res.get("tail_silence") or 0) >= 0.5 else "tail pad mandatory"),
        ("peak / RMS dBFS", "%s / %s" % (audit_res.get("peak_dbfs"), audit_res.get("rms_dbfs")),
         "peak < -1; headroom for music bed", "n/a"),
    ]
    for row in checks:
        au.append(list(row))
    for l in audit_res.get("top_pauses", []):
        au.append(["real pause", l, "visual breather may land here", ""])
    for c in au[1]:
        c.fill = hdr_fill
        c.font = Font(bold=True, color="FFFFFF")
    for r in range(2, au.max_row + 1):
        for c in range(1, 5):
            au.cell(r, c).border = border
            au.cell(r, c).alignment = Alignment(vertical="top", wrap_text=True)
    for col, w in {1: 26, 2: 40, 3: 44, 4: 22}.items():
        au.column_dimensions[get_column_letter(col)].width = w

    wb.save(path)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--transcript", required=True, help="narration.json (final TTS timing)")
    ap.add_argument("--words", help="word-timed transcript .txt")
    ap.add_argument("--beats", help="optional beats JSON (list of [start,end] or {'beats': ...})")
    ap.add_argument("--wav", help="optional narration.wav for level checks")
    ap.add_argument("--out", required=True)
    ap.add_argument("--tail", type=float, default=2.0, help="tail pad seconds (1.5-2.5)")
    args = ap.parse_args()

    segs, dur = load_segments(args.transcript)
    words = load_words(args.words)

    if args.beats:
        beats = load_beats_file(args.beats)
    else:
        beats = derive_beats(segs, words)

    # seam handling: <1.5 s beats get no row of their own; the previous row
    # spans across them (0-s seam segments were already dropped in load_segments)
    rows = []
    seam_notes = {}
    for s, e in beats:
        if e - s < 1.5 and rows:
            prev = rows[-1]
            rows[-1] = (prev[0], max(prev[1], e))
            seam_notes[len(rows)] = True
        elif e - s < 1.5:
            continue  # leading sliver
        else:
            rows.append((s, e))
    if not rows:
        rows = [(0.0, dur)]
    if rows[0][0] > 0.05:
        rows[0] = (0.0, rows[0][1])
    fixed = [rows[0]]
    for r in rows[1:]:
        if r[0] > fixed[-1][1] + 0.05:
            fixed.append((fixed[-1][1], r[0]))
        fixed.append(r)
    rows = fixed
    if rows[-1][1] < dur - 0.05:
        rows.append((rows[-1][1], dur))
    rows.append((dur, dur + args.tail))

    acts = make_acts(rows[:-1])  # tail row is not an act member

    audit_res = audit(segs, dur, words, args.wav)
    build_xlsx(args.out, rows, acts, audit_res, dur, seam_notes)

    # plain summaries in column G (narration text per row, <=60 words)
    import openpyxl as _ox
    wb2 = _ox.load_workbook(args.out)
    wbt = wb2["Timeline"]
    for i, (s, e) in enumerate(rows, start=2):
        t = " ".join(t_ for s_, e_, t_ in segs if s_ < e and e_ > s)
        wbt.cell(i, 7).value = words60(t)
    wb2.save(args.out)

    print(json.dumps(summary, indent=1))


if __name__ == "__main__":
    main()
